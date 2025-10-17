from flask import Flask, render_template, request, redirect, url_for, send_file, Response, jsonify, session
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime
import json
import re
import io
import csv
from typing import List, Dict
import os
import glob
from werkzeug.utils import secure_filename
import openpyxl

# Reference to openpyxl to keep static analyzers quiet (used when exporting xlsx via pandas/openpyxl)
_ = getattr(openpyxl, '__version__', None)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key')

# -------------------- Auth (CSV) --------------------
AUTH_CSV = os.path.join(os.path.dirname(__file__), 'auth.csv')

# Note: Removed automatic creation of auth.csv. The app now expects an existing CSV
# at `AUTH_CSV`. The login flow will attempt to read it and return a helpful
# error if the file is missing.


def load_auth(path: str) -> dict:
    """Load auth CSV into a dict mapping username -> {password, role, department_id}.

    This function will NOT create the file. If the file does not exist it returns an
    empty dict so the caller can handle the condition (for example, by showing an error).
    """
    auth = {}
    if not os.path.exists(path):
        return auth
    try:
        with open(path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            # expect headers: department_id,username,password,role
            if reader.fieldnames is None or 'username' not in reader.fieldnames:
                # fallback: read manually (legacy format)
                f.seek(0)
                for row in csv.reader(f):
                    if not row:
                        continue
                    if row[0] == 'username' or row[0] == 'department_id':
                        continue
                    # legacy: username,password,role
                    user = row[0]
                    pwd = row[1] if len(row) > 1 else ''
                    role = row[2] if len(row) > 2 else 'department'
                    auth[user] = {'password': pwd, 'role': role, 'department_id': ''}
            else:
                for row in reader:
                    user = (row.get('username') or '').strip()
                    if not user:
                        continue
                    pwd = (row.get('password') or '').strip()
                    role = (row.get('role') or 'department').strip()
                    dept_id = (row.get('department_id') or '').strip()
                    auth[user] = {'password': pwd, 'role': role, 'department_id': dept_id}
    except Exception:
        # On any read error, fallback to empty
        return {}
    return auth


def authenticate(username: str, password: str) -> dict:
    """Return auth record if username/password match, else None.

    This function loads the CSV at call time and does not create a file. If the CSV
    is missing or empty the function returns None.
    """
    auth = load_auth(AUTH_CSV)
    if not auth:
        # no auth file or failed to read; treat as authentication failure (login will render helpful message)
        return None
    rec = auth.get(username)
    if not rec:
        return None
    if rec.get('password') == password:
        return {'username': username, 'role': rec.get('role', 'department'), 'department_id': rec.get('department_id', '')}
    return None

# Base link to use when constructing absolute URLs for saved DB links.
# Change this to your deployment base URL when you deploy (e.g. https://example.com)
curr_link = os.environ.get('CURR_LINK', "https://tantra-vl7d.onrender.com")

def make_static_url(filename: str) -> str:
    """Return an absolute URL for a file in the static folder using curr_link as base.

    Example: make_static_url('logos/foo.png') -> 'http://127.0.0.1:5000/static/logos/foo.png'
    """
    # use url_for to get the path portion, then prefix with curr_link
    path = url_for('static', filename=filename, _external=False)
    return curr_link.rstrip('/') + path


def _get_branch(obj: dict) -> str:
    """Return a branch value from a participant/record dict using multiple possible keys.

    Firestore documents in this project have used different field names for branch
    over time (e.g. 'branch/Class', 'branch', 'Class', 'branch_name'). This helper
    centralizes the fallback logic so exports and views consistently include branch.
    """
    if not obj:
        return ''
    return (obj.get('branch/Class') or obj.get('branch') or obj.get('Class') or obj.get('branch_name') or '')


def _normalize_status(val) -> str:
    """Normalize a status value from the database or incoming form to a string 'open' or 'close'.

    Handles numeric (1/0 or '1'/'0'), legacy values 'open'/'closed' and returns 'open' by default.
    """
    if val is None:
        return 'open'
    try:
        # numeric-like
        if isinstance(val, int):
            return 'open' if val == 1 else 'close'
        s = str(val).strip().lower()
        if s in ('1', 'open', 'opened', 'true', 'yes'):
            return 'open'
        if s in ('0', 'close', 'closed', 'false', 'no'):
            return 'close'
    except Exception:
        pass
    return 'open'

# -------------------- Upload folders --------------------
UPLOAD_QR_FOLDER = 'static/qr'
UPLOAD_EVENT_FOLDER = 'static/event_images'
UPLOAD_LOGO_FOLDER = 'static/logos'

os.makedirs(UPLOAD_QR_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_EVENT_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_LOGO_FOLDER, exist_ok=True)

app.config['UPLOAD_QR_FOLDER'] = UPLOAD_QR_FOLDER
app.config['UPLOAD_EVENT_FOLDER'] = UPLOAD_EVENT_FOLDER
app.config['UPLOAD_LOGO_FOLDER'] = UPLOAD_LOGO_FOLDER

# -------------------- Firebase --------------------
# Initialize Firebase using one of the following (in order of preference):
# 1) FIREBASE_SERVICE_ACCOUNT_JSON environment variable containing the
#    service account JSON string
# 2) FIREBASE_CREDENTIALS_FILE environment variable pointing to a file path
# 3) A file checked into the repo (not recommended)

_sa_json = os.environ.get('FIREBASE_SERVICE_ACCOUNT_JSON')
_sa_file = os.environ.get('FIREBASE_CREDENTIALS_FILE', 'techfestadmin-a2e2c-firebase-adminsdk-fbsvc-a2a3aaa0e7.json')

if _sa_json:
    try:
        sa_info = json.loads(_sa_json) if isinstance(_sa_json, str) else _sa_json
        cred = credentials.Certificate(sa_info)
    except Exception as e:
        raise RuntimeError('Failed to parse FIREBASE_SERVICE_ACCOUNT_JSON: ' + str(e))
elif os.path.exists(_sa_file):
    cred = credentials.Certificate(_sa_file)
else:
    # Try to auto-detect a service account file matching the typical filename pattern.
    candidates = glob.glob(os.path.join(os.getcwd(), 'techfestadmin*.json'))
    if candidates:
        # pick the first candidate and proceed
        picked = candidates[0]
        print(f"Using detected Firebase service account file: {picked}")
        cred = credentials.Certificate(picked)
    else:
        raise RuntimeError('Firebase service account not found. Set FIREBASE_SERVICE_ACCOUNT_JSON or FIREBASE_CREDENTIALS_FILE.')

firebase_admin.initialize_app(cred)
db = firestore.client()

# -------------------- Login / Home --------------------


@app.route('/')
def root():
    # initial page should be the login page
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        # ensure auth CSV exists
        if not os.path.exists(AUTH_CSV):
            return render_template('login.html', error=f'Authentication configuration missing: {AUTH_CSV}')

        user = authenticate(username, password)
        if not user:
            return render_template('login.html', error='Invalid credentials')
        # set session
        session['username'] = user['username']
        session['role'] = user['role']
        # store department_id when present
        if user.get('department_id'):
            session['department_id'] = user.get('department_id')
        else:
            session.pop('department_id', None)
        if user['role'] == 'admin':
            return redirect(url_for('index'))
        else:
            return redirect(url_for('department_dashboard'))
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/update_auth')
def update_auth_page():
    """Show the auth update page (admin only)."""
    if session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    try:
        auth_list = []
        # Read the current auth CSV
        if os.path.exists(AUTH_CSV):
            with open(AUTH_CSV, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    auth_list.append({
                        'department_id': row.get('department_id', ''),
                        'username': row.get('username', ''),
                        'password': row.get('password', ''),
                        'role': row.get('role', 'department')
                    })
        
        return render_template('update_auth.html', auth_list=auth_list)
    except Exception as e:
        return str(e), 500


@app.route('/update_auth', methods=['POST'])
def update_auth():
    """Update credentials for a department (admin only)."""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request'}), 400
        
        department_id = data.get('department_id')
        new_username = data.get('username', '').strip()
        new_password = data.get('password', '').strip()
        
        if not department_id or not new_username or not new_password:
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Read current auth data
        auth_data = []
        if os.path.exists(AUTH_CSV):
            with open(AUTH_CSV, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                auth_data = list(reader)
        
        # Update the matching record
        updated = False
        for record in auth_data:
            if record.get('department_id') == department_id:
                record['username'] = new_username
                record['password'] = new_password
                updated = True
                break
        
        if not updated:
            return jsonify({'error': 'Department not found'}), 404
        
        # Write back to CSV
        with open(AUTH_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['department_id', 'username', 'password', 'role'])
            writer.writeheader()
            writer.writerows(auth_data)
        
        return jsonify({'message': 'Credentials updated successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/department')
def department_dashboard():
    # only accessible to department role
    if session.get('role') != 'department':
        return redirect(url_for('login'))
    username = session.get('username')
    department_id = session.get('department_id', username)
    # Resolve department name from Firestore. Try collections 'departments' then 'department'.
    dept_name = None
    try:
        doc = db.collection('departments').document(department_id).get()
        if doc.exists:
            dept_name = doc.to_dict().get('name')
        else:
            # try alternative collection name
            doc2 = db.collection('department').document(department_id).get()
            if doc2.exists:
                dept_name = doc2.to_dict().get('name')
    except Exception:
        # ignore Firestore errors and treat as not found
        dept_name = None

    # If dept_name still not found, fall back to using the department_id as name
    if not dept_name:
        dept_name = department_id

    # Gather participants where participant.department == dept_name
    participants = []
    try:
        parts_q = db.collection('participants').where('department', '==', dept_name).stream()
        for pdoc in parts_q:
            p = pdoc.to_dict()
            participants.append({
                'name': p.get('name', ''),
                'email': p.get('email', ''),
                'phone': p.get('phone', ''),
                'college': p.get('college', ''),
                'branch': _get_branch(p),
                'year': p.get('year', ''),
                'event': p.get('event', ''),
                'transactionId': p.get('transactionId') or p.get('transaction_id') or ''
            })
    except Exception:
        participants = []

    # Sort participants by name
    participants = sorted(participants, key=lambda r: (r.get('name') or '').lower())

    total_participants = len(participants)

    # Unique by participant name (case-insensitive)
    unique_names = set()
    for p in participants:
        n = (p.get('name') or '').strip().lower()
        if n:
            unique_names.add(n)
    unique_count = len(unique_names)

    # Per-event counts (from participants)
    event_counts = {}
    for p in participants:
        en = p.get('event') or ''
        event_counts[en] = event_counts.get(en, 0) + 1

    # Build event list from events collection where department/ dept_id matches department_id
    events_info = []
    try:
        all_events = list(db.collection('events').stream())
        for edoc in all_events:
            ev = edoc.to_dict()
            ev_dept = ev.get('department') or ev.get('dept_id') or ''
            if ev_dept == department_id:
                ev_name = ev.get('name') or ''
                ev_status = _normalize_status(ev.get('status'))
                ev_count = event_counts.get(ev_name, 0)
                events_info.append({'id': edoc.id, 'name': ev_name, 'status': ev_status, 'participant_count': ev_count})
    except Exception:
        events_info = []

    # Also include events that participants reference but which may not have matching event docs
    for ename, cnt in event_counts.items():
        if any(e['name'] == ename for e in events_info):
            continue
        if ename:
            events_info.append({'id': '', 'name': ename, 'status': None, 'participant_count': cnt})    # Sort events_info by name
    events_info = sorted(events_info, key=lambda e: (e.get('name') or '').lower())

    return render_template('fordepartement.html', username=username, department_id=department_id,
                           department_name=dept_name, participants=participants,
                           total_participants=total_participants, unique_count=unique_count,
                           per_event_counts=sorted(event_counts.items(), key=lambda x: x[0].lower()),
                           events_info=events_info)


# -------------------- Home --------------------
@app.route('/index')
def index():
    # Dashboard summary counts
    # Count departments
    depts = list(db.collection('departments').stream())
    total_departments = len(depts)
    # build department name map to avoid repeated lookups
    dept_map = {d.id: d.to_dict().get('name', '') for d in depts}

    # Count events
    events = list(db.collection('events').stream())
    total_events = len(events)

    # Build participant counts per event name (participants store event by name)
    event_counts_map = {}
    try:
        parts_all = list(db.collection('participants').stream())
        for pdoc in parts_all:
            p = pdoc.to_dict()
            ev = (p.get('event') or '').strip()
            if ev:
                event_counts_map[ev] = event_counts_map.get(ev, 0) + 1
    except Exception:
        event_counts_map = {}

    # Count registrations/participants and unique participants
    # Some deployments use a 'registrations' collection; others (your setup) use 'participants'.
    parts = list(db.collection('participants').stream())
    total_registrations = len(parts)
    unique_participants = set()
    for pdoc in parts:
        p = pdoc.to_dict()
        # prefer email as unique id, fallback to phone or doc id
        ident = (p.get('email') or p.get('phone') or pdoc.id)
        if ident:
            unique_participants.add(str(ident).strip().lower())
    total_unique_participants = len(unique_participants)

    # Build a small recent events list for the dashboard
    recent_events = []
    for e in events:
        ed = e.to_dict()
        did = ed.get('department')
        recent_events.append({
            'id': e.id,
            'name': ed.get('name'),
            'dept_id': did,
            # show the department name when known, otherwise show the raw dept id
            'dept_name': dept_map.get(did, (did or '')),
            'date': ed.get('date'),
            'status': _normalize_status(ed.get('status')),
            # participant count (match by event name)
            'participant_count': event_counts_map.get(ed.get('name') or '', 0)
        })

    # prepare a simple departments list for the dashboard (id, name, logo_url)
    dept_list = [(d.id, d.to_dict().get('name', ''), d.to_dict().get('logo_url', '')) for d in depts]

    return render_template('index.html',
                           total_departments=total_departments,
                           total_events=total_events,
                           total_registrations=total_registrations,
                           total_unique_participants=total_unique_participants,
                           recent_events=recent_events,
                           departments=dept_list,
                           username=session.get('username'))





@app.route('/dept_events/<dept_id>', methods=['GET'])
def dept_events(dept_id):
    """Return events for a department as JSON."""
    if not dept_id:
        return jsonify({'events': []})
    try:
        ev_q = db.collection('events').where('department', '==', dept_id).stream()
    except Exception:
        # Fallback: return empty
        return jsonify({'events': []})
    evs = []
    for e in ev_q:
        ed = e.to_dict()
        evs.append({
            'id': e.id,
            'name': ed.get('name'),
            'date': ed.get('date'),
            'status': _normalize_status(ed.get('status')),
            'image_url': ed.get('image_url', ''),
            'venue': ed.get('venue', ''),
            'department': ed.get('department', '')
        })
    return jsonify({'events': evs, 'dept_id': dept_id})


@app.route('/event/<event_id>', methods=['GET'])
def get_event(event_id):
    """Return a single event's details as JSON."""
    if not event_id:
        return jsonify({'error': 'missing id'}), 400
    ev_doc = db.collection('events').document(event_id).get()
    if not ev_doc.exists:
        return jsonify({'error': 'not found'}), 404
    ed = ev_doc.to_dict()
    result = {
        'id': ev_doc.id,
        'name': ed.get('name'),
        'description': ed.get('description', ''),
        'date': ed.get('date', ''),
        'time': ed.get('time', ''),
        'venue': ed.get('venue', ''),
        'image_url': ed.get('image_url', ''),
    'department': ed.get('department', ''),
    'status': _normalize_status(ed.get('status')),
        'price': ed.get('price', ''),
        'prize': ed.get('prize', '')
    }
    return jsonify({'event': result})

# -------------------- Add Department --------------------
@app.route('/add_department', methods=['GET', 'POST'])
def add_department():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']

        # Upload logo
        logo_file = request.files.get('logo_file')
        logo_url = ""
        if logo_file and logo_file.filename != "":
            filename = secure_filename(logo_file.filename)
            path = os.path.join(app.config['UPLOAD_LOGO_FOLDER'], filename)
            logo_file.save(path)
            logo_url = make_static_url(f'logos/{filename}')

        # Upload QR
        qr_file = request.files.get('qr_file')
        qr_url = ""
        if qr_file and qr_file.filename != "":
            filename = secure_filename(qr_file.filename)
            path = os.path.join(app.config['UPLOAD_QR_FOLDER'], filename)
            qr_file.save(path)
            qr_url = make_static_url(f'qr/{filename}')

        db.collection('departments').document().set({
            'name': name,
            'description': description,
            'logo_url': logo_url,
            'qr_url': qr_url,
            'created_at': datetime.utcnow()
        })
        return redirect(url_for('index'))
    # show existing departments on the page for quick reference
    departments = list(db.collection('departments').stream())
    dept_list = [(d.id, d.to_dict().get('name', ''), d.to_dict().get('description', '')) for d in departments]
    return render_template('add_department.html', departments=dept_list)

# -------------------- Add Event --------------------
@app.route('/add_event', methods=['GET', 'POST'])
def add_event():
    departments = db.collection('departments').stream()
    dept_list = [(dept.id, dept.to_dict().get('name', '')) for dept in departments]

    default_date = '2025-10-24'

    if request.method == 'POST':
        dept_id = request.form.get('dept_id')
        name = request.form.get('name', '')
        description = request.form.get('description', '')
        date = request.form.get('date', default_date)
        time = request.form.get('time', '')
        venue = request.form.get('venue', '')

        # Additional fields
        category = request.form.get('category', 'Individual')
        coordinator = request.form.get('coordinator', '')
        coordinatorPhone = request.form.get('coordinatorPhone', '')
        # 'participants' field removed from the form (not stored)

        # Event image upload
        event_file = request.files.get('event_image')
        image_url = ""
        if event_file and event_file.filename != "":
            filename = secure_filename(event_file.filename)
            path = os.path.join(app.config['UPLOAD_EVENT_FOLDER'], filename)
            event_file.save(path)
            image_url = make_static_url(f'event_images/{filename}')

        # Get department QR and store event in top-level `events` collection
        dept_doc = db.collection('departments').document(dept_id).get() if dept_id else None
            # No payment_qr_url needed

        # status: use string 'open' or 'close'. Accept legacy numeric values and normalize.
        status = _normalize_status(request.form.get('status', 'open'))

        price_raw = request.form.get('price', '')
        try:
            price = float(price_raw) if price_raw not in (None, '') else 0
            if isinstance(price, float) and price.is_integer():
                price = int(price)
        except Exception:
            price = price_raw or ''

        prize = request.form.get('prize', '')

        # Find the highest numeric event id in the collection
        all_events = db.collection('events').stream()
        max_id = 0
        for ev in all_events:
            try:
                eid = int(ev.id)
                if eid > max_id:
                    max_id = eid
            except Exception:
                continue
        new_id = max_id + 1
        event_ref = db.collection('events').document(str(new_id))
        event_ref.set({
            'id': new_id,
            'department': dept_id,
            'name': name,
            'description': description,
            'date': date,
            'time': time,
            'venue': venue,
            'image': image_url,
            'image_url': image_url,
            'category': category,
            'coordinator': coordinator,
            'coordinatorPhone': coordinatorPhone,
            'price': price,
            'prize': prize,
            'status': status,
            'created_at': datetime.utcnow()
        })
        return redirect(url_for('add_event'))

    return render_template('add_event.html', departments=dept_list, default_date=default_date)





@app.route('/delete_event', methods=['POST'])
def delete_event():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    event_id = request.form.get('event_id')
    if not event_id:
        return jsonify({'error': 'Event ID is required'}), 400

    try:
        # Get the event to check if it exists and get its department
        event_ref = db.collection('events').document(event_id)
        event = event_ref.get()
        
        if not event.exists:
            return jsonify({'error': 'Event not found'}), 404

        event_data = event.to_dict()
        
        # Check if user has permission to delete this event
        user_dept = session.get('department_id', '')
        user_role = session.get('role', '')
        
        if user_role != 'admin' and user_dept != event_data.get('dept_id'):
            return jsonify({'error': 'Unauthorized to delete this event'}), 403

        # Delete the event
        event_ref.delete()
        
        # Also delete any registrations for this event
        registrations = db.collection('registrations').where('event_id', '==', event_id).stream()
        for reg in registrations:
            reg.reference.delete()

        return jsonify({'message': 'Event deleted successfully'})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/toggle_event_status', methods=['POST'])
def toggle_event_status():
    """Toggle event registration status between 'open' and 'close'.

    Expects form-encoded POST with 'event_id'. Requires logged-in session user.
    Flips string status ('open'->'close' or 'close'->'open') and persists to Firestore.
    Returns JSON with the new status string and boolean is_open.
    """
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    event_id = request.form.get('event_id') or request.json and request.json.get('event_id')
    if not event_id:
        return jsonify({'error': 'Missing event_id'}), 400

    try:
        event_ref = db.collection('events').document(event_id)
        ev = event_ref.get()
        if not ev.exists:
            return jsonify({'error': 'Event not found'}), 404

        data = ev.to_dict() or {}
        current = _normalize_status(data.get('status'))
        new_status = 'close' if current == 'open' else 'open'
        event_ref.update({'status': new_status})

        return jsonify({'success': True, 'status': new_status, 'is_open': new_status == 'open'})

    except Exception as e:
        print('Error toggling event status:', e)
        return jsonify({'error': 'Failed to update event status', 'details': str(e)}), 500


def _resolve_participant_from_registration(reg_data: dict) -> Dict:
    """Given a registration document dict, attempt to resolve participant data.
    Handles keys: participant_id, user_id, 'participant' inline dict, participant_email, email.
    Returns participant dict or None if not found.
    """
    if not reg_data or not isinstance(reg_data, dict):
        return None

    # Inline participant object
    inline = reg_data.get('participant')
    if inline and isinstance(inline, dict):
        return inline

    # Common id fields
    pid = reg_data.get('participant_id') or reg_data.get('user_id') or reg_data.get('uid')
    if pid:
        # Try participants collection first
        ref = db.collection('participants').document(pid).get()
        if ref.exists:
            return ref.to_dict()
        # fallback to users collection
        ref2 = db.collection('users').document(pid).get()
        if ref2.exists:
            return ref2.to_dict()

    # Try to resolve by email if provided in registration
    email = reg_data.get('participant_email') or reg_data.get('email') or reg_data.get('user_email')
    if email:
        # search participants by email
        q = db.collection('participants').where('email', '==', email).limit(1).stream()
        for doc in q:
            return doc.to_dict()
        q2 = db.collection('users').where('email', '==', email).limit(1).stream()
        for doc in q2:
            return doc.to_dict()
    return None

# -------------------- View Participants --------------------
@app.route('/view_participants', methods=['GET'])
def view_participants():
    departments = db.collection('departments').stream()
    dept_list = [(dept.id, dept.to_dict()['name']) for dept in departments]
    # build quick lookup map for department id -> name
    dept_map = {d[0]: d[1] for d in dept_list}

    selected_dept_id = request.args.get('dept_id')
    # template uses 'event_id' (which contains the event name in this dataset)
    selected_event_id = request.args.get('event_id')
    # If an event is selected from the dropdown, enable sorting by event
    sort_by_event = bool(selected_event_id)
    participants_info: List[Dict] = []

    # events_for_select: used to populate events dropdown. Build after we know selected_dept_id

    # Primary source: participants collection (no registrations in your setup)
    parts = db.collection('participants').stream()

    # If a department is selected via dept_id (which is a department document id), translate to department name
    selected_dept_name = None
    if selected_dept_id:
        ddoc = db.collection('departments').document(selected_dept_id).get()
        if ddoc.exists:
            selected_dept_name = ddoc.to_dict().get('name')

    # selected_event_id is taken directly from query string (if any)

    for pdoc in parts:
        p = pdoc.to_dict()
        p_dept = p.get('department') or ''
        p_event = p.get('event') or ''
        # apply filters if provided
        if selected_dept_name and p_dept != selected_dept_name:
            continue
        if selected_event_id and p_event != selected_event_id:
            continue

        participants_info.append({
            'name': p.get('name'),
            'email': p.get('email'),
            'phone': p.get('phone'),
            'college': p.get('college'),
            'branch': _get_branch(p),
            'year': p.get('year'),
            'event_name': p_event,
            'dept_name': p_dept,
            'event_id': '',
            'transaction_id': p.get('transactionId') or p.get('transaction_id')
        })

    # Sort: department name first, then optional event name, then participant name
    if sort_by_event:
        participants_info = sorted(participants_info, key=lambda r: (r.get('dept_name', ''), r.get('event_name', ''), r.get('name', '')))
    else:
        participants_info = sorted(participants_info, key=lambda r: (r.get('dept_name', ''), r.get('name', '')))

    # Build events_for_select now (limit to selected department if provided)
    if selected_dept_id:
        ev_q = db.collection('events').where('department', '==', selected_dept_id).stream()
    else:
        ev_q = db.collection('events').stream()
    events_for_select = [(e.to_dict().get('name'), e.to_dict().get('name')) for e in ev_q]

    return render_template('view_participants.html',
                           departments=dept_list,
                           participants=participants_info,
                           selected_dept_id=selected_dept_id,
                           selected_event_id=selected_event_id,
                           events_for_select=events_for_select)


@app.route('/participants_list', methods=['GET'])
def participants_list():
    """Participants listing for a department with optional event filter and sorting.

    Query params:
      dept_id - department document id (preferred)
      event - event name to filter participants by
      sort - 'event' to sort by event, otherwise sort by name
    """
    dept_id = request.args.get('dept_id')
    selected_event = request.args.get('event')
    sort = request.args.get('sort', '')

    dept_name = None
    events_for_select = []
    # resolve department name from dept_id if provided
    if dept_id:
        try:
            ddoc = db.collection('departments').document(dept_id).get()
            if ddoc.exists:
                dept_name = ddoc.to_dict().get('name')
        except Exception:
            dept_name = None

    # Build event list for this department (by dept id OR by stored department field)
    try:
        ev_q = db.collection('events').stream() if not dept_id else db.collection('events').stream()
        # collect events that match the department id or department field
        evs = []
        for ed in db.collection('events').stream():
            ev = ed.to_dict()
            ev_dept = ev.get('department') or ev.get('dept_id') or ''
            if dept_id:
                if ev_dept == dept_id:
                    evs.append(ev.get('name'))
            else:
                # if no dept_id provided, include all events
                evs.append(ev.get('name'))
        # unique and sorted
        events_for_select = sorted([e for e in set([x for x in evs if x])], key=lambda s: s.lower())
    except Exception:
        events_for_select = []

    # load participants filtered by department name (if available) and by selected event (if provided)
    parts = []
    try:
        q = db.collection('participants')
        if dept_name:
            q = q.where('department', '==', dept_name)
        for pdoc in q.stream():
            p = pdoc.to_dict()
            doc_id = getattr(pdoc, 'id', None)
            pname = p.get('name','')
            pevent = p.get('event','')
            # if an event filter is set, apply it
            if selected_event and (pevent or '').strip() != selected_event:
                continue
            parts.append({
                'name': pname,
                'email': p.get('email',''),
                'phone': p.get('phone',''),
                'college': p.get('college',''),
                'branch': _get_branch(p),
                'year': p.get('year',''),
                'event': pevent,
                'department': p.get('department',''),
                # include the Firestore document id so we can display it if transaction id is missing
                'doc_id': doc_id,
                # robust transaction id lookup (support multiple field names)
                'transaction_id': (
                    p.get('transactionId') or p.get('transaction_id') or p.get('txid') or p.get('transaction') or ''
                )
            })
    except Exception:
        parts = []

    # sorting
    if sort == 'event':
        parts = sorted(parts, key=lambda r: ((r.get('event') or '').lower(), (r.get('name') or '').lower()))
    else:
        parts = sorted(parts, key=lambda r: ((r.get('name') or '').lower()))

    return render_template('participants_list.html', participants=parts, dept_name=dept_name, dept_id=dept_id,
                           events_for_select=events_for_select, selected_event=selected_event, sort=sort)


@app.route('/download_participants')
def download_participants():
    """Download participants list in XLSX or PDF format."""
    if 'username' not in session:
        return redirect(url_for('login'))

    event_name = request.args.get('event')
    format_type = request.args.get('format', 'xlsx')
    dept_id = request.args.get('dept_id')

    # Require either an event or a department filter to avoid exporting the entire database
    if not event_name and not dept_id:
        return Response('Please select a department or an event before downloading.', status=400)

    # Allow exporting all events when event_name is not provided
    participants = []
    try:
        # Build query respecting department and event filters
        q = db.collection('participants')
        if dept_id:
            # resolve department name if possible
            try:
                ddoc = db.collection('departments').document(dept_id).get()
                if ddoc.exists:
                    dept_name = ddoc.to_dict().get('name')
                else:
                    dept_name = dept_id
            except Exception:
                dept_name = dept_id
            q = q.where('department', '==', dept_name)
        if event_name:
            q = q.where('event', '==', event_name)

        for doc in q.stream():
            data = doc.to_dict() or {}
            participants.append({
                'name': data.get('name', ''),
                'event': data.get('event', ''),
                'college': data.get('college', ''),
                'branch': _get_branch(data),
                'year': data.get('year', ''),
                'email': data.get('email', ''),
                'phone': data.get('phone', ''),
                # robust transaction id lookup
                'transaction_id': data.get('transactionId') or data.get('transaction_id') or data.get('txid') or data.get('transaction') or ''
            })
    except Exception as e:
        return str(e), 500

    # Column order should match the page exactly
    export_headers = ['Name', 'Event', 'College', 'Branch', 'Year', 'Email', 'Phone', 'Transaction ID']

    if format_type == 'xlsx':
        try:
            # Create Excel file in memory
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Participants"

            # Write headers
            for col, header in enumerate(export_headers, 1):
                sheet.cell(row=1, column=col, value=header)

            # Write data rows
            for row_idx, p in enumerate(participants, start=2):
                row_vals = [
                    p.get('name', ''),
                    p.get('event', ''),
                    p.get('college', ''),
                    p.get('branch', ''),
                    p.get('year', ''),
                    p.get('email', ''),
                    p.get('phone', ''),
                    p.get('transaction_id', '')
                ]
                for col_idx, val in enumerate(row_vals, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=val)

            # Style the headers
            for cell in sheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            workbook.save(output)
            output.seek(0)

            filename = f"{event_name or 'all_events'}_participants.xlsx"
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            return str(e), 500

    elif format_type == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title=f"{event_name or 'All Events'} - Participants List")

            # Build table data
            table_data = [export_headers]
            for p in participants:
                row = [
                    p.get('name', ''),
                    p.get('event', ''),
                    p.get('college', ''),
                    p.get('branch', ''),
                    p.get('year', ''),
                    p.get('email', ''),
                    p.get('phone', ''),
                    p.get('transaction_id', '')
                ]
                table_data.append(row)

            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a4a4a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ])

            table = Table(table_data)
            table.setStyle(style)

            styles = getSampleStyleSheet()
            heading_style = ParagraphStyle('Heading', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=12)
            heading = Paragraph(f"{event_name or 'All Events'} - Participants List", heading_style)

            doc.build([heading, table])
            buffer.seek(0)
            filename = f"{event_name or 'all_events'}_participants.pdf"
            return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        except Exception as e:
            return str(e), 500

    return "Invalid format type", 400


def _gather_participants(dept_id: str, event_id: str = None) -> List[Dict]:
    """Helper to collect participants for a department, optionally filtering by event id."""
    rows: List[Dict] = []
    if not dept_id:
        return rows

    # Build list of event ids for the department (or single event id if provided)
    event_ids = []
    event_map = {}
    if event_id:
        ev_doc = db.collection('events').document(event_id).get()
        if ev_doc.exists:
            event_ids = [ev_doc.id]
            event_map[ev_doc.id] = ev_doc.to_dict()
    else:
        events = db.collection('events').where('department', '==', dept_id).stream()
        for e in events:
            event_ids.append(e.id)
            event_map[e.id] = e.to_dict()

    if not event_ids:
        return rows

    # Firestore 'in' queries accept up to 10 items; batch if necessary
    BATCH = 10
    for i in range(0, len(event_ids), BATCH):
        batch_ids = event_ids[i:i+BATCH]
        regs_query = db.collection('registrations').where('event_id', 'in', batch_ids).stream()
        for reg in regs_query:
            reg_data = reg.to_dict()
            p = _resolve_participant_from_registration(reg_data)
            if not p:
                continue
            ev_id = reg_data.get('event_id')
            ev_data = event_map.get(ev_id, {})
            rows.append({
                'name': p.get('name'),
                'email': p.get('email'),
                'phone': p.get('phone'),
                'college': p.get('college'),
                'branch': _get_branch(p),
                'year': p.get('year'),
                'event_name': ev_data.get('name'),
                'dept_name': ev_data.get('department') or ev_data.get('dept_id') or '',
                'event_id': ev_id,
                'transaction_id': reg_data.get('transaction_id')
            })

    return rows


@app.route('/export_participants')
def export_participants():
    """Export participants for a department (and optional event) in csv/xlsx/pdf formats.

    Query params: dept_id, event_id (optional), format (csv|xlsx|pdf)
    """
    dept_id = request.args.get('dept_id')
    event_id = request.args.get('event_id')
    fmt = request.args.get('format', 'xlsx').lower()

    # Helper to sanitize filename parts
    def _sanitize(s: str) -> str:
        if not s:
            return ''
        s = s.lower()
        # replace non-alphanumeric with underscore
        s = re.sub(r'[^a-z0-9]+', '_', s)
        s = s.strip('_')
        return s or 'value'

    # Determine dept_name (participants store department by name in this dataset)
    dept_name = None
    if dept_id:
        d = db.collection('departments').document(dept_id).get()
        if d.exists:
            dept_name = d.to_dict().get('name')

    # Determine event_name: try doc id first, else treat as event name string
    event_name = None
    if event_id:
        # try as doc id
        evdoc = db.collection('events').document(event_id).get()
        if evdoc.exists:
            event_name = evdoc.to_dict().get('name')
        else:
            # assume event_id is a name string
            event_name = event_id

    # Build participant rows directly from participants collection (matches view)
    q = db.collection('participants')
    if dept_name:
        q = q.where('department', '==', dept_name)
    if event_name:
        q = q.where('event', '==', event_name)
    rows = []
    for doc in q.stream():
        p = doc.to_dict()
        rows.append({
            'name': p.get('name', ''),
            'email': p.get('email', ''),
            'phone': p.get('phone', ''),
            'college': p.get('college', ''),
            'branch': _get_branch(p),
            'year': p.get('year', ''),
            'event_name': p.get('event', ''),
            'dept_name': p.get('department', ''),
            'transaction_id': p.get('transactionId') or p.get('transaction_id', '')
        })

    # Sort
    rows = sorted(rows, key=lambda r: (r.get('dept_name', ''), r.get('event_name', ''), r.get('name', '')))

    headers = ['name', 'email', 'phone', 'college', 'branch', 'year', 'event_name', 'dept_name', 'transaction_id']

    # Build filename pattern: tantra_{department}_{event}.{ext}
    part_dept = _sanitize(dept_name) if dept_name else 'all_departments'
    part_event = _sanitize(event_name) if event_name else 'all_events'
    base_filename = f'tantra_{part_dept}_{part_event}'

    if fmt == 'xlsx':
        try:
            import pandas as pd
        except Exception:
            return Response('pandas is required to export XLSX. Install with `pip install pandas openpyxl`', status=500)
        df = pd.DataFrame(rows)
        for h in headers:
            if h not in df.columns:
                df[h] = ''
        buf = io.BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        filename = f'{base_filename}.xlsx'
        return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if fmt == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        except Exception:
            return Response('reportlab is required to export PDF. Install with `pip install reportlab`', status=500)
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        data = [headers]
        for r in rows:
            data.append([r.get(h, '') for h in headers])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c4dff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        doc.build([table])
        buf.seek(0)
        filename = f'{base_filename}.pdf'
        return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')

    return Response('Unsupported format. Allowed: xlsx, pdf', status=400)

# -------------------- View Database Content --------------------
@app.route('/db_content')
def db_content():
    departments = db.collection('departments').stream()
    all_data = []
    for dept in departments:
        dept_data = dept.to_dict()
        # load events from top-level collection that belong to this department
        events = db.collection('events').where('dept_id', '==', dept.id).stream()
        event_list = []
        for e in events:
            ev = e.to_dict()
            ev['_id'] = e.id
            event_list.append(ev)
        all_data.append({
            'dept_id': dept.id,
            'dept_name': dept_data.get('name'),
            'description': dept_data.get('description'),
            'logo_url': dept_data.get('logo_url'),
            'qr_url': dept_data.get('qr_url'),
            'events': event_list
        })
    return render_template('db_content.html', data=all_data)


@app.route('/fix_events', methods=['GET', 'POST'])
def fix_events():
    # list departments for selection
    departments = list(db.collection('departments').stream())
    dept_list = [(d.id, d.to_dict().get('name')) for d in departments]

    message = ''
    if request.method == 'POST':
        event_id = request.form.get('event_id')
        new_dept = request.form.get('dept_id')
        if event_id and new_dept:
            db.collection('events').document(event_id).update({'dept_id': new_dept})
            message = 'Updated event department.'

    # find events with missing/unknown dept_id
    events = list(db.collection('events').stream())
    dept_ids = {d.id for d in departments}
    problematic = []
    for e in events:
        ed = e.to_dict()
        did = ed.get('dept_id')
        if not did or did not in dept_ids:
            problematic.append({'id': e.id, 'name': ed.get('name'), 'date': ed.get('date'), 'dept_id': did})

    return render_template('fix_events.html', events=problematic, departments=dept_list, message=message)


@app.route('/repair_events', methods=['GET'])
def repair_events():
    """Admin utility: analyze and optionally repair 'events' documents.

    Query params:
      apply=1  -> apply the safe repairs (otherwise dry-run)

    Repairs performed when apply=1:
      - ensure document contains numeric 'id' field (derived from numeric doc id if possible)
      - if 'dept_id' missing but 'department' present and matches a department doc id, set 'dept_id'
      - ensure 'image' and 'image_url' fields are present when an image_url is available

    Returns a JSON report of findings and applied changes.
    """
    apply = request.args.get('apply', '') == '1'

    # load departments map (id -> name)
    departments = list(db.collection('departments').stream())
    dept_ids = {d.id for d in departments}

    events = list(db.collection('events').stream())
    report = {'total_events': len(events), 'problems': [], 'applied': []}

    for e in events:
        ed = e.to_dict() or {}
        doc_id = e.id
        problems = []
        changes = {}

        # ensure numeric id field
        if not isinstance(ed.get('id'), int):
            try:
                numeric = int(doc_id)
                changes['id'] = numeric
                problems.append('missing_or_nonint_id')
            except Exception:
                # try to parse from existing id field
                try:
                    existing = int(ed.get('id'))
                    changes['id'] = existing
                except Exception:
                    # cannot determine numeric id; skip setting
                    pass

        # ensure dept_id exists: try 'dept_id' then fallback to 'department'
        did = ed.get('dept_id') or ed.get('department')
        if not did or did not in dept_ids:
            problems.append('missing_or_invalid_dept_id')
        else:
            if ed.get('dept_id') != did:
                changes['dept_id'] = did

        # ensure image fields present
        if ed.get('image') and not ed.get('image_url'):
            changes['image_url'] = ed.get('image')
        if ed.get('image_url') and not ed.get('image'):
            changes['image'] = ed.get('image_url')

        if problems:
            report['problems'].append({'doc': doc_id, 'name': ed.get('name'), 'issues': problems, 'proposed_changes': changes})

        if apply and changes:
            try:
                db.collection('events').document(doc_id).update(changes)
                report['applied'].append({'doc': doc_id, 'changes': changes})
            except Exception as ex:
                report.setdefault('errors', []).append({'doc': doc_id, 'error': str(ex)})

    return jsonify(report)

# -------------------- Run --------------------
if __name__ == '__main__':
    # When running locally, allow PORT to be overridden (Render provides $PORT).
    port = int(os.environ.get('PORT', 5000))
    # Bind to 0.0.0.0 so Render (or other hosts) can reach the service.
    app.run(host='0.0.0.0', port=port, debug=True)
